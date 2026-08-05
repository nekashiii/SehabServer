from flask import Flask, render_template, request, redirect, url_for, session, jsonify, send_file, Response
from functools import wraps
import json, os, hashlib, secrets, smtplib, time, queue, threading, pathlib
import requests
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from dotenv import load_dotenv

load_dotenv()  

app = Flask(__name__)
app.secret_key = secrets.token_hex(32)
app.config['TEMPLATES_AUTO_RELOAD'] = True 

BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
USERS_FILE = os.path.join(BASE_DIR, 'users.json')
DADOS_FILE = os.path.join(BASE_DIR, 'dados.json')

SYNC_URL = os.environ['SYNC_URL']
SYNC_INTERVAL_SEGUNDOS = 30 * 60

DENUNCIAS_RECEBIDAS_FILE = os.path.join(BASE_DIR, 'denuncias_recebidas.txt')

def get_denuncias_recebidas():
    """Le o numero de denuncias recebidas do arquivo denuncias_recebidas.txt.
    Se o arquivo nao existir ou tiver algo invalido, cria/mantem um padrao
    e avisa no console, sem derrubar o servidor."""
    padrao = 6578
    try:
        with open(DENUNCIAS_RECEBIDAS_FILE, encoding='utf-8') as f:
            texto = f.read().strip()
        return int(texto)
    except FileNotFoundError:
        with open(DENUNCIAS_RECEBIDAS_FILE, 'w', encoding='utf-8') as f:
            f.write(str(padrao))
        print(f"[denuncias] Arquivo {DENUNCIAS_RECEBIDAS_FILE} nao existia, criado com {padrao}.")
        return padrao
    except ValueError:
        print(f"[denuncias] Conteudo invalido em {DENUNCIAS_RECEBIDAS_FILE}, usando {padrao}.")
        return padrao

PROG_DATA_DIR    = pathlib.Path(BASE_DIR) / 'data'
PROG_UPLOAD_PATH = PROG_DATA_DIR / 'planilha_atual.xlsx'
PROG_DATA_DIR.mkdir(exist_ok=True)

_clients: list[queue.Queue] = []
_clients_lock = threading.Lock()

def _broadcast(event: str, data: dict):
    msg = f"event: {event}\ndata: {json.dumps(data)}\n\n"
    with _clients_lock:
        dead = []
        for q in _clients:
            try:
                q.put_nowait(msg)
            except queue.Full:
                dead.append(q)
        for q in dead:
            _clients.remove(q)


def sincronizar_fiscalizacao():
    """Busca os dados calculados no Power Automate (Planilha Secretario +
    Sistematizacao) e atualiza o dados.json, preservando denuncias/repetidas
    (que continuam manuais)."""
    try:
        resp = requests.get(SYNC_URL, timeout=30)
        resp.raise_for_status()
        novos_dados = resp.json()
    except Exception as e:
        print(f'[sync] Falha ao buscar dados online: {e}')
        return

    dados_atuais = {}
    if os.path.exists(DADOS_FILE):
        try:
            with open(DADOS_FILE, encoding='utf-8') as f:
                dados_atuais = json.load(f)
        except Exception:
            dados_atuais = {}

    # Preserva os campos manuais
    for campo_manual in ('denuncias', 'repetidas'):
        if campo_manual in dados_atuais:
            novos_dados.setdefault(campo_manual, dados_atuais[campo_manual])

    # 'denuncias' vem sempre do arquivo denuncias_recebidas.txt,
    # não é editável pelo dashboard.
    novos_dados['denuncias'] = get_denuncias_recebidas()

    # 'repetidas' é sempre recalculado: denuncias - notificados (nunca negativo)
    notificados = novos_dados.get('notificados', 0) or 0
    denuncias   = novos_dados.get('denuncias', 0) or 0
    novos_dados['repetidas'] = max(0, denuncias - notificados)

    novos_dados['atualizado_por'] = 'Sincronizacao automatica'
    novos_dados['atualizado_em']  = time.strftime('%d/%m/%Y %H:%M')

    with open(DADOS_FILE, 'w', encoding='utf-8') as f:
        json.dump(novos_dados, f, ensure_ascii=False)

    print(f"[sync] dados.json atualizado as {novos_dados['atualizado_em']}")
    _broadcast('nova_planilha', {
        'msg': f"Dados sincronizados automaticamente as {novos_dados['atualizado_em']}"
    })


def _loop_sincronizacao():
    while True:
        sincronizar_fiscalizacao()
        time.sleep(SYNC_INTERVAL_SEGUNDOS)


SENHA_PADRAO    = os.environ['SENHA_PADRAO']
EMAIL_REMETENTE = os.environ['EMAIL_REMETENTE']
EMAIL_SENHA_APP = os.environ['EMAIL_SENHA_APP']
EMAIL_SMTP      = os.environ.get('EMAIL_SMTP', 'smtp.gmail.com')
EMAIL_PORTA     = int(os.environ.get('EMAIL_PORTA', '587'))
SERVER_BASE_URL = os.environ['SERVER_BASE_URL']

reset_tokens = {}
def load_users():
    if not os.path.exists(USERS_FILE):
        return {}
    with open(USERS_FILE, encoding='utf-8') as f:
        return json.load(f)

def save_users(users):
    with open(USERS_FILE, 'w', encoding='utf-8') as f:
        json.dump(users, f, indent=2, ensure_ascii=False)

def hash_senha(senha):
    return hashlib.sha256(senha.encode()).hexdigest()

def buscar_por_usuario(usuario, users):
    u = usuario.strip().lower()
    for email, data in users.items():
        if data.get('matricula', '').lower() == u:
            return email, data
    return None, None

def enviar_email(destinatario, assunto, corpo):
    msg = MIMEMultipart()
    msg['From']    = EMAIL_REMETENTE
    msg['To']      = destinatario
    msg['Subject'] = assunto
    msg.attach(MIMEText(corpo, 'plain', 'utf-8'))
    with smtplib.SMTP(EMAIL_SMTP, EMAIL_PORTA, timeout=10) as srv:
        srv.starttls()
        srv.login(EMAIL_REMETENTE, EMAIL_SENHA_APP)
        srv.send_message(msg)
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'usuario' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def api_login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'usuario' not in session:
            return jsonify({'erro': 'Nao autenticado'}), 401
        return f(*args, **kwargs)
    return decorated

@app.route('/', methods=['GET', 'POST'])
@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'usuario' in session:
        return redirect(url_for('selecionar'))
    erro = None
    msg  = request.args.get('msg')
    if request.method == 'POST':
        usuario_input = request.form.get('usuario', '').strip()
        senha         = request.form.get('senha', '').strip()
        users         = load_users()
        user_email, user = buscar_por_usuario(usuario_input, users)
        if user_email is None:
            erro = 'Usuario nao encontrado. Verifique e tente novamente.'
        elif user.get('bloqueado'):
            erro = 'Conta bloqueada apos muitas tentativas. Fale com o Lucas.'
        else:
            tentativas = user.get('tentativas', 0)
            if user['senha'] == hash_senha(senha):
                users[user_email]['tentativas'] = 0
                save_users(users)
                session['usuario'] = user_email
                session['nome']    = user.get('nome', user_email)
                if user.get('primeiro_acesso', True):
                    return redirect(url_for('trocar_senha'))
                return redirect(url_for('selecionar'))
            else:
                tentativas += 1
                users[user_email]['tentativas'] = tentativas
                if tentativas >= 5:
                    users[user_email]['bloqueado'] = True
                    save_users(users)
                    erro = 'Conta bloqueada apos 5 tentativas incorretas. Fale com o Lucas.'
                else:
                    save_users(users)
                    restantes = 5 - tentativas
                    erro = f'Senha incorreta. Voce ainda tem {restantes} tentativa(s) antes do bloqueio.'
    return render_template('login.html', erro=erro, msg=msg)


@app.route('/esqueci-senha', methods=['GET', 'POST'])
def esqueci_senha():
    msg  = None
    erro = None
    if request.method == 'POST':
        usuario_input    = request.form.get('usuario', '').strip()
        users            = load_users()
        user_email, user = buscar_por_usuario(usuario_input, users)
        if user_email is None:
            erro = 'Usuario nao encontrado. Verifique o seu user.'
        else:
            token = secrets.token_urlsafe(32)
            reset_tokens[token] = {'email': user_email, 'expira': time.time() + 1800}
            link = f"{SERVER_BASE_URL}/redefinir-senha/{token}"
            nome = user.get('nome', usuario_input)
            corpo = (
                f'Ola, {nome}!\n\n'
                'Recebemos uma solicitacao para redefinir sua senha do Painel SEHAB.\n\n'
                'Clique no link abaixo para criar uma nova senha (valido por 30 minutos):\n\n'
                f'{link}\n\n'
                'Se voce nao solicitou isso, ignore este e-mail.\n\n'
                'Atenciosamente,\nSEHAB - Fiscalizacao HIS/HMP\nPrefeitura de Sao Paulo'
            )
            try:
                enviar_email(user_email, 'SEHAB - Redefinicao de senha', corpo)
                msg = f'Link enviado para o e-mail cadastrado de {nome}.'
            except Exception as e:
                erro = f'Erro ao enviar e-mail. Fale com o Lucas. ({e})'
    return render_template('esqueci_senha.html', msg=msg, erro=erro)


@app.route('/redefinir-senha/<token>', methods=['GET', 'POST'])
def redefinir_senha(token):
    info = reset_tokens.get(token)
    if not info or time.time() > info['expira']:
        return render_template('esqueci_senha.html', erro='Link expirado ou invalido. Solicite um novo.', msg=None)
    erro = None
    if request.method == 'POST':
        nova     = request.form.get('nova_senha', '').strip()
        confirma = request.form.get('confirma_senha', '').strip()
        if len(nova) < 6:
            erro = 'A senha deve ter pelo menos 6 caracteres.'
        elif nova != confirma:
            erro = 'As senhas nao coincidem. Tente novamente.'
        elif nova == SENHA_PADRAO:
            erro = 'Escolha uma senha diferente da senha padrao.'
        else:
            users = load_users()
            email = info['email']
            users[email]['senha']           = hash_senha(nova)
            users[email]['primeiro_acesso'] = False
            users[email]['tentativas']      = 0
            users[email]['bloqueado']       = False
            save_users(users)
            del reset_tokens[token]
            return redirect(url_for('login', msg='Senha redefinida com sucesso! Faca login.'))
    return render_template('trocar_senha.html', erro=erro, nome='', via_token=True, token=token)


@app.route('/trocar-senha', methods=['GET', 'POST'])
@login_required
def trocar_senha():
    erro = None
    if request.method == 'POST':
        nova     = request.form.get('nova_senha', '').strip()
        confirma = request.form.get('confirma_senha', '').strip()
        if len(nova) < 6:
            erro = 'A senha deve ter pelo menos 6 caracteres.'
        elif nova != confirma:
            erro = 'As senhas nao coincidem. Tente novamente.'
        elif nova == SENHA_PADRAO:
            erro = 'Escolha uma senha diferente da senha padrao.'
        else:
            users = load_users()
            email = session['usuario']
            users[email]['senha']           = hash_senha(nova)
            users[email]['primeiro_acesso'] = False
            users[email]['tentativas']      = 0
            save_users(users)
            return redirect(url_for('selecionar'))
    return render_template('trocar_senha.html', erro=erro, nome=session.get('nome'), via_token=False)


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


@app.route('/admin/desbloquear/<matricula>')
@login_required
def desbloquear(matricula):
    users = load_users()
    email, user = buscar_por_usuario(matricula, users)
    if email:
        users[email]['bloqueado']  = False
        users[email]['tentativas'] = 0
        save_users(users)
        return f'Usuario {user["nome"]} desbloqueado com sucesso. <a href="/">Voltar</a>'
    return 'Usuario nao encontrado. <a href="/">Voltar</a>'


@app.route('/selecionar')
@login_required
def selecionar():
    nome = session.get('nome', session.get('usuario'))
    return render_template('selecionar.html', nome=nome)

@app.route('/dashboard')
@login_required
def dashboard():
    users = load_users()
    email = session['usuario']
    nome  = users[email].get('nome', email)
    return render_template('dashboard.html', nome=nome, email=email)


@app.route('/notificacao')
@login_required
def notificacao():
    return render_template('notificação.html')

@app.route('/dados/carregar', methods=['GET'])
@api_login_required
def dados_carregar():
    if not os.path.exists(DADOS_FILE):
        return jsonify({}), 200
    with open(DADOS_FILE, encoding='utf-8') as f:
        return jsonify(json.load(f))


@app.route('/prog/')
@login_required
def prog_dashboard():
    return render_template('prog_dashboard.html')

@app.route('/prog/upload', methods=['POST'])
@login_required
def prog_upload():
    if 'file' not in request.files:
        return jsonify({'ok': False, 'msg': 'Nenhum arquivo enviado.'}), 400
    f = request.files['file']
    if not f.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({'ok': False, 'msg': 'Arquivo deve ser .xlsx ou .xls.'}), 400
    f.save(PROG_UPLOAD_PATH)
    ts = time.strftime('%d/%m/%Y %H:%M:%S')
    usuario = session.get('nome', 'desconhecido')
    print(f"[{ts}] Nova planilha recebida por {usuario}: {f.filename} ({PROG_UPLOAD_PATH.stat().st_size // 1024} KB)")
    _broadcast('nova_planilha', {
        'nome': f.filename,
        'ts':   ts,
        'msg':  f'Planilha atualizada as {ts}'
    })
    return jsonify({'ok': True, 'msg': f'Planilha recebida as {ts}'})

@app.route('/prog/planilha')
@login_required
def prog_download_planilha():
    if not PROG_UPLOAD_PATH.exists():
        return 'Nenhuma planilha carregada ainda.', 404
    return send_file(
        PROG_UPLOAD_PATH,
        as_attachment=False,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/prog/status')
@login_required
def prog_status():
    if PROG_UPLOAD_PATH.exists():
        mtime = PROG_UPLOAD_PATH.stat().st_mtime
        ts = time.strftime('%d/%m/%Y %H:%M:%S', time.localtime(mtime))
        return jsonify({'disponivel': True, 'atualizada_em': ts})
    return jsonify({'disponivel': False})

@app.route('/prog/events')
def prog_events():
    # NAO usa @login_required: redirect(302) quebra EventSource no browser
    if 'usuario' not in session:
        return Response('Nao autenticado', status=401)
    q: queue.Queue = queue.Queue(maxsize=10)
    with _clients_lock:
        _clients.append(q)

    def stream():
        yield "event: ping\ndata: {}\n\n"
        try:
            while True:
                try:
                    msg = q.get(timeout=25)
                    yield msg
                except queue.Empty:
                    yield "event: ping\ndata: {}\n\n"
        except GeneratorExit:
            pass
        finally:
            with _clients_lock:
                if q in _clients:
                    _clients.remove(q)

    return Response(
        stream(),
        mimetype='text/event-stream',
        headers={
            'Cache-Control':     'no-cache',
            'X-Accel-Buffering': 'no',
            'Connection':        'keep-alive',
        }
    )



@app.route('/lote')
@login_required
def lote_page():
    return render_template('lote.html')


@app.route('/extrator')
@login_required
def extrator_page():
    return render_template('extrator.html')


import re as _re_ext
import pdfplumber as _pdfplumber
import logging as _logging
from openpyxl import load_workbook as _load_workbook

_logging.getLogger('pdfminer').setLevel(_logging.ERROR)

_progresso_extrator = {}
_output_extrator    = {}

try:
    from ftfy import fix_text as _fix_text
except ImportError:
    def _fix_text(t): return t  # fallback gracioso se ftfy não instalado


def _limpar_texto_ext(texto):
    texto = _fix_text(texto)
    texto = texto.replace('\xa0', ' ')
    texto = _re_ext.sub(r'\s+', ' ', texto)
    return texto.strip()


def _corrigir_nome(nome):
    partes = nome.split()
    resultado = []
    i = 0
    while i < len(partes):
        if len(partes[i]) == 1 and i + 1 < len(partes):
            resultado.append(partes[i] + partes[i+1])
            i += 2
        else:
            resultado.append(partes[i])
            i += 1
    return ' '.join(resultado)


def _formatar_data(data_texto):
    meses = {
        'janeiro':'01','fevereiro':'02','março':'03','marco':'03',
        'abril':'04','maio':'05','junho':'06','julho':'07',
        'agosto':'08','setembro':'09','outubro':'10',
        'novembro':'11','dezembro':'12'
    }
    partes = data_texto.lower().split(' de ')
    if len(partes) != 3:
        return data_texto
    dia = partes[0].zfill(2)
    mes = meses.get(partes[1], '??')
    ano = partes[2]
    return f'{dia}/{mes}/{ano}'


def _extrair_compradores(texto):
    bloco = _re_ext.search(r'adquirentes:\s*(.*)', texto, _re_ext.IGNORECASE)
    if not bloco:
        return 'NÃO ENCONTRADO'
    trecho = bloco.group(1)
    padrao = r'([A-ZÁÉÍÓÚÂÊÔÃÕÇ\s]+?),\s*(?:nacionalidade|brasileir)'
    nomes = _re_ext.findall(padrao, trecho, _re_ext.IGNORECASE)
    nomes_limpos = []
    for nome in nomes:
        nome = _limpar_texto_ext(nome)
        nome = _corrigir_nome(nome)
        if len(nome) > 3:
            nomes_limpos.append(nome)
    return ', '.join(nomes_limpos) if nomes_limpos else 'NÃO ENCONTRADO'


def _encontrar_comprador(pdf):
    for pagina in pdf.pages[:4]:
        try:
            texto = pagina.extract_text()
        except Exception:
            continue
        if not texto:
            continue
        texto = _limpar_texto_ext(texto)
        if 'adquirentes' in texto.lower():
            return _extrair_compradores(texto)
    compradores = []
    for i in range(25, 40):
        if i >= len(pdf.pages):
            continue
        try:
            texto = pdf.pages[i].extract_text()
        except Exception:
            continue
        if not texto:
            continue
        texto = _limpar_texto_ext(texto)
        padrao = r'([A-ZÁÉÍÓÚÂÊÔÃÕÇ\s]+)\s+CPF:.*?Assinou como parte compradora'
        nomes = _re_ext.findall(padrao, texto, _re_ext.IGNORECASE | _re_ext.DOTALL)
        for nome in nomes:
            nome = _limpar_texto_ext(nome)
            nome = _corrigir_nome(nome)
            if len(nome) > 3:
                compradores.append(nome)
    if compradores:
        return ', '.join(compradores)
    compradores = []
    for i in range(4, 9):
        if i >= len(pdf.pages):
            continue
        try:
            texto = pdf.pages[i].extract_text()
        except Exception:
            continue
        if not texto:
            continue
        texto = _limpar_texto_ext(texto)
        padrao = r'Cliente:\s*([A-ZÁÉÍÓÚÂÊÔÃÕÇ\s]+)'
        nomes = _re_ext.findall(padrao, texto, _re_ext.IGNORECASE)
        for nome in nomes:
            nome = _limpar_texto_ext(nome)
            nome = _corrigir_nome(nome)
            if len(nome) > 3:
                compradores.append(nome)
    return ', '.join(compradores) if compradores else 'NÃO ENCONTRADO'


def _extrair_data_assinatura(pdf):
    padrao = r'São Paulo,\s*(\d{1,2}\s+de\s+[A-Za-zçÇãõéÉ]+\s+de\s+\d{4})'
    for i in range(8, 14):
        if i >= len(pdf.pages):
            continue
        try:
            texto = pdf.pages[i].extract_text()
        except Exception:
            continue
        if not texto:
            continue
        texto = _limpar_texto_ext(texto)
        resultado = _re_ext.search(padrao, texto, _re_ext.IGNORECASE)
        if resultado:
            return _formatar_data(resultado.group(1))
    padrao2 = r'Assinou como parte compradora em (\d{1,2}\s+[a-zA-Z]{3}\s+\d{4})'
    for i in range(25, 40):
        if i >= len(pdf.pages):
            continue
        try:
            texto = pdf.pages[i].extract_text()
        except Exception:
            continue
        if not texto:
            continue
        texto = _limpar_texto_ext(texto)
        resultado = _re_ext.search(padrao2, texto, _re_ext.IGNORECASE)
        if resultado:
            return _formatar_data(resultado.group(1))
    padrao3 = r'Data Base:\s*(\d{2}/\d{2}/\d{4})'
    for i in range(4, 9):
        if i >= len(pdf.pages):
            continue
        try:
            texto = pdf.pages[i].extract_text()
        except Exception:
            continue
        if not texto:
            continue
        texto = _limpar_texto_ext(texto)
        resultado = _re_ext.search(padrao3, texto, _re_ext.IGNORECASE)
        if resultado:
            return resultado.group(1)
    return 'NÃO ENCONTRADO'


def _processar_zip_extrator(zip_path, process_id):
    try:
        temp_dir = _tempfile.mkdtemp()
        _apagar_pasta_delay(temp_dir, 10)
        with _zipfile.ZipFile(zip_path, 'r') as zip_ref:
            zip_ref.extractall(temp_dir)
        pasta_docs = _os.path.join(temp_dir, 'documentos')
        if not _os.path.exists(pasta_docs):
            raise Exception("Pasta 'documentos' não encontrada no ZIP")
        arquivos = [f for f in _os.listdir(pasta_docs) if f.lower().endswith('.pdf')]
        dados = []
        _progresso_extrator[process_id].update({'total': len(arquivos), 'atual': 0, 'status': 'processando'})
        for i, arquivo in enumerate(arquivos):
            caminho_pdf = _os.path.join(pasta_docs, arquivo)
            numeros = _re_ext.findall(r'\d+', arquivo)
            if not numeros:
                _progresso_extrator[process_id]['atual'] = i + 1
                continue
            unidade = int(numeros[-1])
            try:
                with _pdfplumber.open(caminho_pdf) as pdf:
                    comprador = _encontrar_comprador(pdf)
                    data      = _extrair_data_assinatura(pdf)
                dados.append((unidade, comprador, data))
            except Exception as exc:
                print('ERRO PDF extrator:', exc)
                dados.append((unidade, 'ERRO', 'ERRO'))
            _progresso_extrator[process_id]['atual'] = i + 1
        dados.sort(key=lambda x: x[0])
        base_path = _os.path.join(BASE_DIR, 'data', 'modelo_base.xlsx')
        if not _os.path.exists(base_path):
            raise Exception('modelo_base.xlsx não encontrado em data/')
        wb = _load_workbook(base_path)
        ws = wb.active
        linha_inicio = 2
        for i, (unidade, comprador, data) in enumerate(dados):
            linha = linha_inicio + i
            ws.cell(row=linha, column=1, value=unidade)
            ws.cell(row=linha, column=4, value=comprador)
            ws.cell(row=linha, column=8, value=data)
        output_file = _os.path.join(temp_dir, 'planilha_preenchida.xlsx')
        wb.save(output_file)
        _output_extrator[process_id]             = output_file
        _progresso_extrator[process_id]['status'] = 'finalizado'
    except Exception as exc:
        print('ERRO EXTRATOR:', exc)
        _progresso_extrator[process_id]['status'] = 'erro'


@app.route('/extrator/upload', methods=['POST'])
@login_required
def extrator_upload():
    process_id = str(_uuid.uuid4())
    _progresso_extrator[process_id] = {'atual': 0, 'total': 0, 'status': 'processando'}
    file      = request.files['file']
    pasta_tmp = _tempfile.mkdtemp()
    zip_path  = _os.path.join(pasta_tmp, 'entrada.zip')
    file.save(zip_path)
    _threading.Thread(target=_processar_zip_extrator, args=(zip_path, process_id), daemon=True).start()
    return jsonify({'id': process_id})


@app.route('/extrator/progresso/<process_id>')
@login_required
def extrator_progresso(process_id):
    retorno = dict(_progresso_extrator.get(process_id, {}))
    retorno['download'] = (retorno.get('status') == 'finalizado' and process_id in _output_extrator)
    return jsonify(retorno)


@app.route('/extrator/download/<process_id>')
@login_required
def extrator_download(process_id):
    if process_id not in _output_extrator:
        return 'Arquivo não gerado', 404
    path = _output_extrator[process_id]
    response = _send_file(path, as_attachment=True)
    response.call_on_close(lambda: _shutil.rmtree(_os.path.dirname(path), ignore_errors=True))
    return response




import os as _os
import re as _re
import tempfile as _tempfile
import shutil as _shutil
import zipfile as _zipfile
import threading as _threading
import unicodedata as _unicodedata
import uuid as _uuid
from flask import send_file as _send_file
from pypdf import PdfReader as _PdfReader, PdfWriter as _PdfWriter

_progresso_lote = {}
_output_lote    = {}


def _apagar_pasta_delay(pasta, delay=10):
    def _rm():
        try:
            _shutil.rmtree(pasta, ignore_errors=True)
        except Exception:
            pass
    t = _threading.Timer(delay, _rm)
    t.daemon = True
    t.start()


def _normalizar(texto):
    texto = texto.lower()
    return ''.join(c for c in _unicodedata.normalize('NFD', texto)
                   if _unicodedata.category(c) != 'Mn')


def _agrupar_por_unidade(base_dir):
    grupos = {}
    for root, _, files in _os.walk(base_dir):
        for file in files:
            if file.lower().endswith('.pdf'):
                caminho = _os.path.join(root, file)
                pasta   = _os.path.basename(_os.path.dirname(caminho))
                grupos.setdefault(pasta, []).append(caminho)
    return grupos


def _separar_matriculas(arquivos):
    matriculas, documentos = [], []
    for arq in arquivos:
        nome  = _normalizar(_os.path.basename(arq))
        pasta = _normalizar(_os.path.dirname(arq))
        if 'matricula' in nome or 'matricula' in pasta:
            matriculas.append(arq)
        else:
            documentos.append(arq)
    return matriculas, documentos


def _unir_pdfs(arquivos, log):
    writer  = _PdfWriter()
    sucesso = False
    for arq in sorted(arquivos):
        try:
            reader = _PdfReader(arq)
            if reader.is_encrypted:
                reader.decrypt('')
            for page in reader.pages:
                writer.add_page(page)
            sucesso = True
        except Exception as exc:
            log['erros'].append(f'{arq} -> {exc}')
    return writer if sucesso else None


def _processar_lote(upload_dir, process_id):
    grupos = _agrupar_por_unidade(upload_dir)
    _progresso_lote[process_id].update({'total': len(grupos), 'atual': 0, 'status': 'processando'})

    from datetime import datetime as _dt
    timestamp  = _dt.now().strftime('%Y%m%d_%H%M%S')
    output_tmp = _tempfile.mkdtemp()
    _apagar_pasta_delay(output_tmp, 10)

    zip_path = _os.path.join(output_tmp, f'PDFs_Unificados_{timestamp}.zip')
    log = {'sucesso': 0, 'erros': [], 'sem_matricula': set()}

    with _zipfile.ZipFile(zip_path, 'w', _zipfile.ZIP_DEFLATED) as zipf:
        for unidade, arquivos in grupos.items():
            matriculas, documentos = _separar_matriculas(arquivos)
            if not matriculas:
                log['sem_matricula'].add(unidade)
            if matriculas:
                w = _unir_pdfs(matriculas, log)
                if w:
                    nome = f'matricula_{unidade}.pdf'
                    path = _os.path.join(output_tmp, nome)
                    with open(path, 'wb') as f:
                        w.write(f)
                    zipf.write(path, f'matriculas/{nome}')
                    log['sucesso'] += 1
            if documentos:
                w = _unir_pdfs(documentos, log)
                if w:
                    nome = f'documento_{unidade}.pdf'
                    path = _os.path.join(output_tmp, nome)
                    with open(path, 'wb') as f:
                        w.write(f)
                    zipf.write(path, f'documentos/{nome}')
                    log['sucesso'] += 1
            _progresso_lote[process_id]['atual'] += 1

    _output_lote[process_id]             = zip_path
    _progresso_lote[process_id]['status'] = 'finalizado'


@app.route('/lote/upload', methods=['POST'])
@login_required
def lote_upload():
    process_id = request.form.get('id')
    upload_dir = request.form.get('dir')
    if not process_id:
        process_id = str(_uuid.uuid4())
        upload_dir = _tempfile.mkdtemp()
        _progresso_lote[process_id] = {'total': 0, 'atual': 0, 'status': 'idle'}
    for file in request.files.getlist('files'):
        caminho = file.filename.replace('\\', '/')
        destino = _os.path.join(upload_dir, caminho)
        _os.makedirs(_os.path.dirname(destino), exist_ok=True)
        file.save(destino)
    return jsonify({'status': 'ok', 'dir': upload_dir, 'id': process_id})


@app.route('/lote/processar', methods=['POST'])
@login_required
def lote_processar():
    data       = request.json
    upload_dir = data['dir']
    process_id = data['id']
    _threading.Thread(target=_processar_lote, args=(upload_dir, process_id), daemon=True).start()
    return jsonify({'status': 'processando'})


@app.route('/lote/progresso/<process_id>')
@login_required
def lote_progresso(process_id):
    retorno = dict(_progresso_lote.get(process_id, {}))
    retorno['download'] = (retorno.get('status') == 'finalizado' and process_id in _output_lote)
    return jsonify(retorno)


@app.route('/lote/download/<process_id>')
@login_required
def lote_download(process_id):
    if process_id not in _output_lote:
        return 'Arquivo não gerado', 404
    path = _output_lote[process_id]
    response = _send_file(path, as_attachment=True)
    def _cleanup(r):
        try:
            _shutil.rmtree(_os.path.dirname(path), ignore_errors=True)
        except Exception:
            pass
        return r
    response.call_on_close(lambda: _cleanup(response))
    return response


@app.route('/lote/limpar', methods=['POST'])
@login_required
def lote_limpar():
    temp_dir = _tempfile.gettempdir()
    removidos = 0
    for nome in _os.listdir(temp_dir):
        if nome.lower().startswith('tmp'):
            caminho = _os.path.join(temp_dir, nome)
            try:
                if _os.path.isdir(caminho):
                    _shutil.rmtree(caminho, ignore_errors=True)
                    removidos += 1
            except Exception:
                pass
    return jsonify({'removidos': removidos})



if __name__ == '__main__':
    import socket

    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(('8.8.8.8', 80))
        local_ip = s.getsockname()[0]
        s.close()
    except Exception:
        local_ip = '127.0.0.1'

    print('=' * 60)
    print('  SEHAB - Servidor Unificado')
    print('  Dashboard Fiscalizacao + Dashboard PROG')
    print('=' * 60)
    print(f'  Local:   http://localhost:5000')
    print(f'  Rede:    http://{local_ip}:5000')
    print(f'  Pasta:   {BASE_DIR}')
    print('=' * 60)
    print('  Ctrl+C para encerrar\n')

    threading.Thread(target=_loop_sincronizacao, daemon=True).start()

porta = 5000
app.run(host='0.0.0.0', port=porta, debug=False, threaded=True)

